from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.utils.safestring import mark_safe
import pandas as pd
import openpyxl

from .models import ExcelFile, QueryLog


@admin.register(ExcelFile)
class ExcelFileAdmin(admin.ModelAdmin):
    list_display = ['name', 'file_size_display', 'sheet_count_display', 'uploaded_at', 'is_active', 'view_file_link']
    list_filter = ['is_active', 'uploaded_at']
    search_fields = ['name', 'description']
    readonly_fields = ['uploaded_at', 'updated_at', 'sheet_names', 'column_info', 'file_preview']

    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'description', 'file', 'is_active')
        }),
        ('File Information', {
            'fields': ('uploaded_at', 'updated_at', 'sheet_names', 'column_info'),
            'classes': ('collapse',)
        }),
        ('File Preview', {
            'fields': ('file_preview',),
            'classes': ('collapse',)
        }),
    )

    def file_size_display(self, obj):
        return obj.get_file_size()
    file_size_display.short_description = 'File Size'

    def sheet_count_display(self, obj):
        return obj.get_sheet_count()
    sheet_count_display.short_description = 'Sheets'

    def view_file_link(self, obj):
        if obj.file:
            return format_html('<a href="{}" target="_blank">View File</a>', obj.file.url)
        return "No file"
    view_file_link.short_description = 'File Link'

    def file_preview(self, obj):
        """Show a preview of the Excel file contents"""
        if not obj.file:
            return "No file uploaded"

        try:
            # Get basic file info
            wb = openpyxl.load_workbook(obj.file.path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            preview_html = f"<h4>File: {obj.file.name}</h4>"
            preview_html += f"<p><strong>Sheets ({len(sheet_names)}):</strong> {', '.join(sheet_names)}</p>"

            # Preview first sheet
            if sheet_names:
                try:
                    df = pd.read_excel(obj.file.path, sheet_name=sheet_names[0], nrows=5)
                    preview_html += f"<h5>Preview of '{sheet_names[0]}' (first 5 rows):</h5>"
                    preview_html += "<div style='overflow-x: auto;'>"
                    preview_html += df.to_html(classes='table table-striped table-sm', table_id=None)
                    preview_html += "</div>"

                    # Column info
                    preview_html += f"<p><strong>Total Columns:</strong> {len(df.columns)}</p>"
                    preview_html += f"<p><strong>Columns:</strong> {', '.join(df.columns.tolist())}</p>"

                except Exception as e:
                    preview_html += f"<p style='color: red;'>Error reading sheet: {str(e)}</p>"

            return mark_safe(preview_html)

        except Exception as e:
            return f"Error reading file: {str(e)}"

    file_preview.short_description = 'File Preview'

    def save_model(self, request, obj, form, change):
        """Override save to process Excel file metadata"""
        super().save_model(request, obj, form, change)

        # Process file to extract metadata
        if obj.file:
            try:
                wb = openpyxl.load_workbook(obj.file.path, read_only=True)
                obj.sheet_names = wb.sheetnames

                # Get column information for each sheet
                column_info = {}
                for sheet_name in wb.sheetnames:
                    try:
                        df = pd.read_excel(obj.file.path, sheet_name=sheet_name, nrows=0)
                        column_info[sheet_name] = {
                            'columns': df.columns.tolist(),
                            'column_count': len(df.columns)
                        }
                    except:
                        continue

                obj.column_info = column_info
                wb.close()
                obj.save()

            except Exception as e:
                pass  # Handle silently in admin


@admin.register(QueryLog)
class QueryLogAdmin(admin.ModelAdmin):
    list_display = ['excel_file', 'sheet_name', 'result_found', 'query_time', 'filters_preview']
    list_filter = ['result_found', 'query_time', 'excel_file']
    search_fields = ['excel_file__name', 'sheet_name']
    readonly_fields = ['excel_file', 'sheet_name', 'filters_applied', 'result_found', 'result_data', 'query_time']
    date_hierarchy = 'query_time'

    def filters_preview(self, obj):
        """Show a preview of applied filters"""
        if obj.filters_applied:
            filters = []
            for key, value in obj.filters_applied.items():
                filters.append(f"{key}={value}")
            return ", ".join(filters[:3]) + ("..." if len(filters) > 3 else "")
        return "No filters"
    filters_preview.short_description = 'Applied Filters'

    def has_add_permission(self, request):
        """Disable manual addition of query logs"""
        return False

    def has_change_permission(self, request, obj=None):
        """Make query logs read-only"""
        return False


# Customize admin site
admin.site.site_header = "Excel Analyzer Admin"
admin.site.site_title = "Excel Analyzer"
admin.site.index_title = "Welcome to Excel Analyzer Administration"
