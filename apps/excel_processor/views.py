from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.conf import settings
from django.core.files.storage import default_storage
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
import pandas as pd
import openpyxl
import json
import numpy as np

from .models import ExcelFile, QueryLog, CustomUser

# Test commit
def is_admin(user):
    return user.is_staff

def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            if user.is_staff:
                return redirect('excel_processor:admin_panel')
            return redirect('excel_processor:index')
        else:
            messages.error(request, 'Invalid username or password')
    
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('excel_processor:admin_panel')
        return redirect('excel_processor:index')
    
    return render(request, 'excel_processor/login.html')

@login_required
def logout_view(request):
    logout(request)
    return redirect('excel_processor:login')

@login_required
@user_passes_test(is_admin)
def admin_panel(request):
    context = {
        'users': CustomUser.objects.filter(is_staff=False),
        'excel_files': ExcelFile.objects.all(),
        'search_logs': QueryLog.objects.all().order_by('-query_time')[:100]  # Get last 100 logs
    }
    return render(request, 'excel_processor/admin_panel.html', context)

@login_required
@user_passes_test(is_admin)
def create_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists')
        else:
            CustomUser.objects.create_user(username=username, password=password)
            messages.success(request, 'User created successfully')
        
        return redirect('excel_processor:admin_panel')
    return redirect('excel_processor:admin_panel')

@login_required
@user_passes_test(is_admin)
@require_POST
def toggle_user(request):
    user_id = request.POST.get('user_id')
    user = CustomUser.objects.get(id=user_id)
    user.is_active = not user.is_active
    user.save()
    return JsonResponse({'status': 'success'})

@login_required
@user_passes_test(is_admin)
@require_POST
def upload_excel(request):
    name = request.POST.get('name')
    file = request.FILES.get('file')
    
    if ExcelFile.objects.filter(name=name).exists():
        messages.error(request, 'File name already exists')
    else:
        excel_file = ExcelFile.objects.create(
            name=name,
            file=file,
            uploaded_by=request.user
        )
        
        # Read the Excel file to get sheet names and columns
        try:
            df = pd.ExcelFile(excel_file.file.path)
            sheet_names = df.sheet_names
            
            # Initialize sheet configuration
            sheet_config = {}
            for sheet in sheet_names:
                df_sheet = pd.read_excel(excel_file.file.path, sheet_name=sheet)
                columns = list(df_sheet.columns)
                sheet_config[sheet] = {
                    'is_enabled': False,  # Default to disabled
                    'filter_columns': [],  # Default to no filter columns
                    'result_columns': ['total']  # Default to just 'total'
                }
            
            # Update the model with sheet information and initial configuration
            excel_file.sheet_names = sheet_names
            excel_file.sheet_config = sheet_config
            excel_file.enabled_sheets = list(sheet_names)  # Initially enable all sheets
            excel_file.save()
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
            excel_file.delete()
            return redirect('excel_processor:admin_panel')
        
        messages.success(request, 'File uploaded successfully')
    
    return redirect('excel_processor:admin_panel')

@login_required
@user_passes_test(is_admin)
def configure_sheets(request, excel_id):
    excel_file = get_object_or_404(ExcelFile, id=excel_id)
    
    if request.method == 'POST':
        sheet_config = excel_file.sheet_config or {}
        sheet_name = request.POST.get('sheet_name')
        is_enabled = request.POST.get('is_enabled') == 'true'
        filter_columns = request.POST.getlist('filter_columns[]')
        result_columns = request.POST.getlist('result_columns[]')
        
        # Initialize sheet configuration if it doesn't exist
        if sheet_name not in sheet_config:
            sheet_config[sheet_name] = {}
            
        # Update sheet configuration
        sheet_config[sheet_name] = {
            'is_enabled': is_enabled,
            'filter_columns': filter_columns,
            'result_columns': result_columns if 'total' in result_columns else ['total'] + result_columns
        }
        
        # Update both sheet_config and enabled_sheets
        excel_file.sheet_config = sheet_config
        excel_file.enabled_sheets = [
            sheet for sheet, config in sheet_config.items() 
            if config.get('is_enabled', True)
        ]
        excel_file.save()
        
        return JsonResponse({'status': 'success'})
    
    # Get available columns for each sheet
    sheet_columns = {}
    sheet_config = excel_file.sheet_config or {}
    
    try:
        df = pd.ExcelFile(excel_file.file.path)
        for sheet in excel_file.sheet_names:
            df_sheet = pd.read_excel(excel_file.file.path, sheet_name=sheet)
            sheet_columns[sheet] = list(df_sheet.columns)
            
            # Initialize sheet config if it doesn't exist
            if sheet not in sheet_config:
                sheet_config[sheet] = {
                    'is_enabled': True,  # Default to enabled
                    'filter_columns': [],  # Default to no filter columns
                    'result_columns': ['total']  # Default to just total
                }
        
        # Update the model with initial configuration
        excel_file.sheet_config = sheet_config
        excel_file.enabled_sheets = [
            sheet for sheet, config in sheet_config.items() 
            if config.get('is_enabled', True)
        ]
        excel_file.save()
        
    except Exception as e:
        messages.error(request, f'Error reading Excel file: {str(e)}')
        return redirect('excel_processor:admin_panel')
    
    context = {
        'excel_file': excel_file,
        'sheet_columns': sheet_columns
    }
    return render(request, 'excel_processor/configure_sheets.html', context)

    return redirect('excel_processor:admin_panel')

@login_required
@user_passes_test(is_admin)
@require_POST
def toggle_excel(request):
    excel_id = request.POST.get('excel_id')
    excel = ExcelFile.objects.get(id=excel_id)
    excel.is_active = not excel.is_active
    excel.save()
    return JsonResponse({'status': 'success'})

@login_required
@user_passes_test(is_admin)
@require_POST
def delete_excel(request):
    excel_id = request.POST.get('excel_id')
    excel = ExcelFile.objects.get(id=excel_id)
    if excel.file:
        default_storage.delete(excel.file.name)
    excel.delete()
    return JsonResponse({'status': 'success'})

@login_required
@user_passes_test(is_admin)
def configure_columns(request, excel_id):
    excel_file = get_object_or_404(ExcelFile, id=excel_id)
    
    # Ensure sheet names are cached
    if not excel_file.sheet_names:
        wb = openpyxl.load_workbook(excel_file.file.path, read_only=True)
        excel_file.sheet_names = wb.sheetnames
        wb.close()
        excel_file.save()
    
    if request.method == 'POST':
        # Get enabled sheets
        enabled_sheets = request.POST.getlist('enabled_sheets')
        excel_file.enabled_sheets = enabled_sheets
        
        # Initialize sheet configuration
        sheet_config = {}
        
        # Read all sheets to get their columns
        all_sheets_df = pd.read_excel(excel_file.file.path, sheet_name=None)
        
        for sheet_name in excel_file.sheet_names:
            if sheet_name in enabled_sheets:
                # Get columns for this sheet
                df = all_sheets_df[sheet_name]
                all_columns = df.columns.tolist()
                
                # Get selected columns for this sheet
                filter_columns = request.POST.getlist(f'filter_columns_{sheet_name}')
                result_columns = request.POST.getlist(f'result_columns_{sheet_name}')
                
                # Validate that selected columns exist in the file
                filter_columns = [col for col in filter_columns if col in all_columns]
                result_columns = [col for col in result_columns if col in all_columns]
                
                # Ensure 'total' is always first in result columns if present
                if 'total' in all_columns and 'total' not in result_columns:
                    result_columns.insert(0, 'total')
                
                # Store configuration for this sheet
                sheet_config[sheet_name] = {
                    'filter_columns': filter_columns,
                    'result_columns': result_columns
                }
        
        # Update the model
        excel_file.sheet_config = sheet_config
        excel_file.save()
        
        messages.success(request, 'Sheet and column configuration updated successfully')
        return redirect('excel_processor:admin_panel')
    
    # Read all sheets and their columns for the form
    all_sheets_df = pd.read_excel(excel_file.file.path, sheet_name=None)
    sheet_columns = {sheet: df.columns.tolist() for sheet, df in all_sheets_df.items()}
    
    context = {
        'excel_file': excel_file,
        'all_sheets': excel_file.sheet_names,
        'enabled_sheets': excel_file.enabled_sheets or excel_file.sheet_names,  # Default all enabled
        'sheet_columns': sheet_columns,
        'sheet_config': excel_file.sheet_config or {}  # Default empty config
    }
    return render(request, 'excel_processor/configure_columns.html', context)

@login_required
def index(request):
    """Main page with dynamic dropdowns"""
    excel_files = ExcelFile.objects.filter(is_active=True)
    context = {
        'excel_files': excel_files,
        'result_columns': [col for col in settings.RESULT_COLUMNS if col.lower() != 'total'],
    }

    return render(request, 'excel_processor/index.html', context)


@require_GET
def get_sheets(request):
    """AJAX endpoint to get sheets for selected Excel file"""
    try:
        file_id = request.GET.get('file_id')
        if not file_id:
            return JsonResponse({'error': 'File ID is required'}, status=400)

        excel_file = get_object_or_404(ExcelFile, id=file_id, is_active=True)

        # If sheet names are not cached, read from file
        if not excel_file.sheet_names:
            try:
                wb = openpyxl.load_workbook(excel_file.file.path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()

                # Update the model with sheet names
                excel_file.sheet_names = sheet_names
                excel_file.save()
            except Exception as e:
                return JsonResponse({'error': f'Error reading Excel file: {str(e)}'}, status=500)

        # Filter to only enabled sheets from sheet_config
        enabled_sheets = []
        sheet_config = excel_file.sheet_config or {}
        for sheet in excel_file.sheet_names:
            if sheet_config.get(sheet, {}).get('is_enabled', True):  # Default to True if not configured
                enabled_sheets.append(sheet)

        return JsonResponse({
            'sheets': enabled_sheets,
            'file_name': excel_file.name
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_GET  
def get_columns(request):
    """AJAX endpoint to get filterable columns for selected sheet"""
    try:
        file_id = request.GET.get('file_id')
        sheet_name = request.GET.get('sheet_name')

        if not file_id or not sheet_name:
            return JsonResponse({'error': 'File ID and sheet name are required'}, status=400)

        excel_file = get_object_or_404(ExcelFile, id=file_id, is_active=True)

        # Get sheet configuration
        sheet_config = excel_file.sheet_config.get(sheet_name, {})
        
        # Check if sheet is enabled in sheet_config
        if not sheet_config.get('is_enabled', True):  # Default to True if not configured
            return JsonResponse({'error': 'Selected sheet is not enabled'}, status=400)
        
        # Read the Excel file and get columns
        try:
            df = pd.read_excel(excel_file.file.path, sheet_name=sheet_name)
            all_columns = df.columns.tolist()

            # Get configured filter columns
            filterable_columns = sheet_config.get('filter_columns', [])

            # Get unique values for each filterable column
            column_data = {}
            for column in filterable_columns:
                # Get unique values and convert to string, handling NaN values
                unique_values = df[column].dropna().astype(str).unique().tolist()
                # Sort the values
                unique_values.sort()
                column_data[column] = unique_values

            # Get result columns from sheet config
            result_columns = sheet_config.get('result_columns', ['total'])

            return JsonResponse({
                'columns': column_data,
                'result_columns': result_columns
            })

        except Exception as e:
            return JsonResponse({'error': f'Error reading sheet: {str(e)}'}, status=500)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
@csrf_exempt
def fetch_results(request):
    """AJAX endpoint to fetch results based on selected filters"""
    try:
        data = json.loads(request.body)
        file_id = data.get('file_id')
        sheet_name = data.get('sheet_name')
        filters = data.get('filters', {})

        if not file_id or not sheet_name:
            return JsonResponse({'error': 'File ID and sheet name are required'}, status=400)

        excel_file = get_object_or_404(ExcelFile, id=file_id, is_active=True)

        # Get sheet configuration
        sheet_config = excel_file.sheet_config.get(sheet_name, {})
        
        # Check if sheet is enabled in sheet_config
        if not sheet_config.get('is_enabled', True):  # Default to True if not configured
            return JsonResponse({'error': 'Selected sheet is not enabled'}, status=400)

        try:
            # Read the Excel file
            df = pd.read_excel(excel_file.file.path, sheet_name=sheet_name)

            # Apply filters
            filtered_df = df.copy()
            applied_filters = {}

            for column, value in filters.items():
                if value and column in df.columns:
                    # Convert both to string for comparison to handle mixed types
                    filtered_df = filtered_df[filtered_df[column].astype(str) == str(value)]
                    applied_filters[column] = value

            if len(filtered_df) > 0:
                # Get result columns from sheet config
                sheet_config = excel_file.sheet_config.get(sheet_name, {})
                result_columns = sheet_config.get('result_columns', ['total'])

                # Convert numpy values to native Python types
                results = {}
                
                # First add 'total' if it's in the result columns
                if 'total' in result_columns:
                    total_value = filtered_df['total'].iloc[0]
                    # Convert to float for consistent decimal handling
                    if pd.isna(total_value):
                        total_value = None
                    else:
                        # Convert any numeric type to float
                        try:
                            total_value = float(total_value)
                        except:
                            total_value = None
                    results['total'] = total_value
                
                # Then add all other columns
                for col in result_columns:
                    if col.lower() != 'total':  # Skip total as it's already added
                        value = filtered_df[col].iloc[0]
                        # Convert numpy types to native Python types
                        if isinstance(value, (np.int64, np.int32)):
                            value = int(value)
                        elif isinstance(value, (np.float64, np.float32)):
                            value = float(value)
                        elif pd.isna(value):
                            value = None
                        else:
                            value = str(value)
                        results[col] = value

                # Log the successful search
                QueryLog.objects.create(
                    user=request.user,
                    excel_file=excel_file,
                    sheet_name=sheet_name,
                    filters_applied=applied_filters,
                    result_found=True,
                    result_data=results
                )
                return JsonResponse({
                    'success': True,
                    'results': results,
                    'message': 'Results found successfully!'
                })

            # Log unsuccessful search
            QueryLog.objects.create(
                user=request.user,
                excel_file=excel_file,
                sheet_name=sheet_name,
                filters_applied=applied_filters,
                result_found=False,
                result_data=None
            )
            
            return JsonResponse({
                'success': False,
                'message': 'No results found for the selected filters.',
                'applied_filters': applied_filters
            })

        except Exception as e:
            return JsonResponse({'error': f'Error reading or processing file: {str(e)}'}, status=500)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


def analytics(request):
    """Analytics page showing query statistics"""
    total_queries = QueryLog.objects.count()
    successful_queries = QueryLog.objects.filter(result_found=True).count()
    recent_queries = QueryLog.objects.select_related('excel_file')[:10]

    # Popular files
    from django.db.models import Count
    popular_files = ExcelFile.objects.annotate(
        query_count=Count('querylog')
    ).order_by('-query_count')[:5]

    context = {
        'total_queries': total_queries,
        'successful_queries': successful_queries,
        'success_rate': (successful_queries / total_queries * 100) if total_queries > 0 else 0,
        'recent_queries': recent_queries,
        'popular_files': popular_files,
    }
    return render(request, 'excel_processor/analytics.html', context)
